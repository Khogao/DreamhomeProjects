import os
import re
import sys
from pathlib import Path
import pandas as pd
import numpy as np

"""
Enhanced converter: reads a Markdown file, splits it into sections by H2 (##) headings,
extracts exact "Sheet X: Name" titles as sheet names, converts Markdown tables to
DataFrames with numeric formatting, and writes each section to a separate Excel sheet.

Usage:
    python convert_md_to_xlsx.py "Phan tich.md"

Output:
    Phan tich.xlsx (same folder as input)
"""

HEADING_RE = re.compile(r'^(#{1,6})\s*(.+)$')
SHEET_TITLE_RE = re.compile(r'^Sheet\s+\d+:\s*(.+)$', re.IGNORECASE)
TABLE_SEPARATOR_RE = re.compile(r'^\s*\|?\s*:?[-]+[:\s\|]+$')
TABLE_ROW_RE = re.compile(r'^\s*\|')
CURRENCY_KEYWORDS = ['triệu', 'tỷ', 'vnđ', 'vnd', 'đồng', 'tiền', 'thành tiền', 'doanh thu', 'chi phí', 'lợi nhuận', 'nợ', 'thu', 'trđ']
NUMERIC_KEYWORDS = ['số lượng', 'diện tích', 'm2', 'm²', 'unit', 'căn', 'tầng', 'số', 'lượng', 'tỷ lệ', '%']


def parse_markdown_sections(text):
    lines = text.splitlines()
    sections = []
    current = {'title': 'Intro', 'level': 0, 'lines': []}

    for line in lines:
        m = HEADING_RE.match(line)
        if m:
            level = len(m.group(1))
            title = m.group(2).strip()
            # Start new section
            sections.append(current)
            current = {'title': title, 'level': level, 'lines': []}
        else:
            current['lines'].append(line)
    sections.append(current)
    # drop any empty leading placeholder
    if sections and sections[0]['title'] == 'Intro' and not any(s.strip() for s in sections[0]['lines']):
        sections = sections[1:]
    return sections


def extract_tables_from_lines(lines):
    """Return list of tables (as list of list-of-rows) and remaining text blocks."""
    tables = []
    blocks = []
    buf = []
    in_table = False

    for line in lines:
        if TABLE_ROW_RE.match(line):
            buf.append(line.strip())
            in_table = True
        else:
            if in_table:
                # flush table
                tables.append(buf[:])
                buf = []
                in_table = False
            blocks.append(line)
    if in_table and buf:
        tables.append(buf)
    return tables, '\n'.join(b for b in blocks if b.strip())


def is_numeric_column(header):
    """Check if column header suggests numeric data"""
    header_lower = header.lower()
    return any(keyword in header_lower for keyword in NUMERIC_KEYWORDS)

def is_currency_column(header):
    """Check if column header suggests currency data"""
    header_lower = header.lower()
    return any(keyword in header_lower for keyword in CURRENCY_KEYWORDS)

def clean_numeric_value(value):
    """Clean and convert numeric values, handling Vietnamese number formats"""
    if not isinstance(value, str):
        return value
    
    # Remove common non-numeric characters but keep decimal separators
    cleaned = re.sub(r'[^\d.,\-]', '', value.replace(',', '.'))
    
    # Handle empty strings
    if not cleaned or cleaned in ['-', '.', '']:
        return np.nan
    
    try:
        # Convert to float
        return float(cleaned)
    except (ValueError, TypeError):
        return value  # Return original if can't convert

def md_table_to_dataframe(table_lines):
    # remove leading/trailing pipes and split by |
    rows = []
    for ln in table_lines:
        # remove leading/trailing | and spaces
        row = [c.strip() for c in re.split(r'\|', ln.strip().strip('|'))]
        rows.append(row)
    
    if len(rows) >= 2 and all(re.match(r'^:?-+:?$', h) for h in rows[1]):
        headers = rows[0]
        data = rows[2:]
    else:
        headers = [f'Col{i+1}' for i in range(len(rows[0]))]
        data = rows
    
    # normalize row lengths
    maxc = max(len(headers), max((len(r) for r in data), default=0))
    hdr = headers + [''] * (maxc - len(headers))
    norm = [r + [''] * (maxc - len(r)) for r in data]
    
    df = pd.DataFrame(norm, columns=hdr)
    
    # Convert numeric columns
    for col in df.columns:
        if col and (is_numeric_column(col) or is_currency_column(col)):
            df[col] = df[col].apply(clean_numeric_value)
    
    return df


def extract_sheet_name(title):
    """Extract clean sheet name from 'Sheet X: Name' format"""
    match = SHEET_TITLE_RE.match(title.strip())
    if match:
        return match.group(1).strip()
    return title

def safe_sheet_name(name):
    # Excel sheet name max 31 chars, cannot contain \ / * ? [ ] :
    name = re.sub(r'[\\/*?:\[\]]', '_', name)
    name = name[:31]
    if not name:
        return 'Sheet'
    return name

def format_worksheet(worksheet, df):
    """Apply formatting to Excel worksheet for better readability"""
    from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Create header style
    try:
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
    except ValueError:
        # Style already exists
        header_style = worksheet.parent.named_styles["header_style"]
    
    # Apply header formatting
    for col_num, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.style = header_style
    
    # Auto-adjust column widths and format numeric columns
    for col_num, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        
        # Calculate column width
        max_length = len(str(column)) + 2
        for row_num in range(2, len(df) + 2):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)) + 2)
        
        # Set column width (max 50 chars)
        worksheet.column_dimensions[col_letter].width = min(max_length, 50)
        
        # Format numeric columns
        if is_currency_column(column):
            # Format as currency (Vietnamese style)
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    cell.number_format = '#,##0'
        elif is_numeric_column(column):
            # Format as number
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    cell.number_format = '#,##0.##'


def main():
    if len(sys.argv) < 2:
        print('Usage: python convert_md_to_xlsx.py "Phan tich.md"')
        sys.exit(1)
    md_path = Path(sys.argv[1])
    if not md_path.exists():
        print(f'File not found: {md_path}')
        sys.exit(1)

    text = md_path.read_text(encoding='utf-8')
    sections = parse_markdown_sections(text)

    out_path = md_path.with_suffix('.xlsx')

    writer = pd.ExcelWriter(out_path, engine='openpyxl')

    for sec in sections:
        title = sec['title']
        # Extract clean sheet name from "Sheet X: Name" format
        clean_title = extract_sheet_name(title)
        sheet_name = safe_sheet_name(clean_title)
        
        # extract tables
        tables, remaining = extract_tables_from_lines(sec['lines'])
        if tables:
            for i, tbl in enumerate(tables):
                try:
                    df = md_table_to_dataframe(tbl)
                except Exception as e:
                    print(f"Warning: Failed to parse table in section '{title}': {e}")
                    # fallback: write raw
                    df = pd.DataFrame({'Content': ['\n'.join(tbl)]})
                
                sub_sheet = sheet_name if i == 0 else f"{sheet_name}_{i+1}"
                sub_sheet = safe_sheet_name(sub_sheet)
                df.to_excel(writer, sheet_name=sub_sheet, index=False)
                
                # Apply formatting to the worksheet
                worksheet = writer.sheets[sub_sheet]
                format_worksheet(worksheet, df)
                
        if remaining:
            # write remaining text into a single-column sheet
            # if a sheet with same name exists, append _text
            text_sheet = safe_sheet_name(sheet_name if not tables else f"{sheet_name}_text")
            pd.DataFrame({'Content': remaining.splitlines()}).to_excel(writer, sheet_name=text_sheet, index=False)

    writer.close()
    print(f'Wrote Excel file: {out_path}')


if __name__ == '__main__':
    main()
